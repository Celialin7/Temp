

"""
============================================================
SBERT-Powered Call Coverage Checker (Mandarin Analysis)
============================================================

This script is an advanced version of the call coverage checker, specifically
optimized for Mandarin (Simplified Chinese) transcripts. It replaces the traditional
TF-IDF algorithm with a powerful Sentence-BERT (SBERT) model for superior
semantic similarity understanding.

Key Upgrades:
1. **Core Engine: Sentence-BERT (SBERT):** Uses 'paraphrase-multilingual-MiniLM-L12-v2'
   for deep semantic understanding, moving from keyword matching to intent matching.
2. **Mandarin Tokenizer: Jieba:** Replaces the Cantonese-specific 'pycantonese'
   with 'jieba', the industry standard for Mandarin word segmentation.
3. **Streamlined & Focused:** All TF-IDF related logic has been removed, resulting
   in a cleaner and more modern codebase.
4. **Enhanced Pattern Recognition:** Supports date/numeric pattern recognition for scoring boosts.
5. **SBERT优化的简化验证策略:** 采用"简化的单向验证"而非复杂的双重打分：
   - 信任SBERT的语义理解能力，以"单句匹配"为主
   - 通过放宽上下文窗口(350字符)为SBERT提供丰富上下文
   - 让对话组本身扮演"整段"角色，无需复杂的双重验证
6. **3-Pass Speaker Grouping:** 优化的3-pass分组策略，为SBERT提供最佳语义环境。
7. **Configurable Input Files:** Supports separate call text and script files.

============================================================
"""

# =============================
# Imports & Dependencies
# =============================
import pandas as pd
import jieba  # Replaced pycantonese with jieba for Mandarin
import re
import os
import sys
from difflib import SequenceMatcher
import numpy as np
import tracemalloc
import time

# --- SBERT related imports ---
# Ensure 'sentence-transformers' is installed in your environment
try:
    from sentence_transformers import SentenceTransformer, util
except ImportError:
    print("❌ Error: 'sentence-transformers' package not found.")
    print("Please install it in your environment using the offline method.")
    sys.exit(1)

# Ensure proper encoding for Windows
if sys.platform.startswith('win'):
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.detach())
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.detach())

# Import unified dictionaries (supports both Mandarin and Cantonese)
try:
    import sys
    sys.path.append('..')  # Add parent directory to path
    import dictionaries
except ImportError:
    print("❌ Error: dictionaries.py not found. Please ensure it exists in the project root.")
    sys.exit(1)

# =============================
# USER CONFIGURATION SECTION
# =============================
"""
CONFIGURE YOUR DISCUSSION POINTS FOR ENHANCED SCORING HERE:
============================================================

Instructions:
1. Add discussion point names that you want to enhance with date/numeric pattern recognition
2. Set 'date_boost': score boost (0.05-0.25) for points that should get bonus for date patterns
3. Set 'numeric_boost': score boost (0.05-0.25) for points that should get bonus for numeric patterns
4. Set both to 0.0 if you don't want that type of enhancement

Examples:
- For points about dates/deadlines: set date_boost > 0, numeric_boost = 0
- For points about prices/amounts: set numeric_boost > 0, date_boost = 0
- For mixed points: set both > 0
"""

# MODIFY THIS DICTIONARY TO CONFIGURE YOUR DISCUSSION POINTS:
USER_ENHANCED_DISCUSSION_POINTS = {
    # Example entries - modify these as needed for Mandarin:
    '价格确认': {'date_boost': 0.0, 'numeric_boost': 0.20},
    '风险评估': {'date_boost': 0.15, 'numeric_boost': 0.0},
    '交易价格确认': {'date_boost': 0.0, 'numeric_boost': 0.18},
    '交易数量确认': {'date_boost': 0.0, 'numeric_boost': 0.20},
    '产品信息确认': {'date_boost': 0.0, 'numeric_boost': 0.12},
    '重要事项声明': {'date_boost': 0.12, 'numeric_boost': 0.0},
    '财务状况评估': {'date_boost': 0.0, 'numeric_boost': 0.15},
    
    # Template for adding more:
    # 'Your Discussion Point Name': {'date_boost': 0.1, 'numeric_boost': 0.15},
}

# =============================
# INPUT FILE CONFIGURATION
# =============================
"""
CONFIGURE YOUR INPUT FILE PATHS HERE:
====================================

Instructions:
1. Set CALL_TEXT_FILE_PATH to the path of your call transcript file
2. Set SCRIPT_FILE_PATH to the path of your script/required points file
3. Use forward slashes (/) or raw strings for Windows paths
4. Can be relative paths (from script location) or absolute paths

File Format Requirements:
- Call text file: Must contain columns 'Speaker Roles', 'Transcription', 'Segment Start Time', 'Segment End Time'
- Script file: Must contain columns 'Required_Discussion_Point', 'Standard_Script'
"""

# ==========================================================
# CONFIGURATION (Paths are now passed via function parameters)
# ==========================================================
# Note: File paths and sheet names are now provided via run_analysis() function parameters
# when called from run_batch_analysis.py or other batch processing systems.

# =============================
# SBERT MODEL CONFIGURATION
# =============================
"""
SBERT MODEL PATH CONFIGURATION:
===============================

Set the path to your local SBERT model folder.
The model should be the 'paraphrase-multilingual-MiniLM-L12-v2' model.
"""

# SBERT Model path - loaded from config.py
# Import config.py for SBERT model path
try:
    import sys
    sys.path.insert(0, '..')
    from config import SBERT_MODEL_PATH
except ImportError:
    # Fallback path only when config.py is not available
    SBERT_MODEL_PATH = '../multilingual_sbert/paraphrase-multilingual-MiniLM-L12-v2'
    print("⚠️  Warning: config.py not found, using fallback SBERT model path")

# =============================
# UTILITY FUNCTIONS
# =============================
def print_checkpoint(step_num, description, start_time=None):
    """Progress tracking function"""
    current_time = time.time()
    timestamp = time.strftime('%H:%M:%S')
    if start_time is not None:
        elapsed = current_time - start_time
        print(f"✅ CHECKPOINT {step_num}: {description} (Completed at: {timestamp}, Duration: {elapsed:.2f}s)")
    else:
        print(f"🚀 CHECKPOINT {step_num}: {description} (Started at: {timestamp})")
    return current_time

# ==========================================================
# Class: SbertCallCoverageChecker
# ==========================================================
class SbertCallCoverageChecker:
    def __init__(self, model_path):
        """
        Initialization: Loads dictionaries and the SBERT model.
        """
        # Load dictionaries (now from shared dictionaries.py)
        import sys
        sys.path.insert(0, '..')  # Add parent directory to path
        import dictionaries as shared_dict
        
        self.mandarin_synonyms = shared_dict.cantonese_synonyms  # Use shared synonyms
        self.error_patterns = shared_dict.error_patterns
        self.important_keywords = shared_dict.important_keywords
        # Language-specific stopwords will be loaded in load_call_specific_weights
        self.stopwords = None  # Will be set based on detected language
        
        # Multi-product weight management
        self.current_language = None
        self.current_product = None
        self.current_weights = {}  # Will be loaded by load_call_specific_weights
        
        # Performance optimization caches
        self._preprocessed_texts_cache = {}
        self._similarity_cache = {}
        self._reverse_synonym_map = None

        # --- Load the SBERT model ---
        if not os.path.exists(model_path):
            print(f"❌ FATAL ERROR: SBERT model path does not exist: '{model_path}'")
            print("Please update the SBERT_MODEL_PATH variable in the script.")
            sys.exit(1)
        
        try:
            print(f"Loading SBERT model from: {model_path} ... (This may take a moment)")
            self.sbert_model = SentenceTransformer(model_path)
            print("✅ SBERT model loaded successfully.")
        except Exception as e:
            print(f"❌ FATAL ERROR: Failed to load SBERT model from '{model_path}'.")
            print(f"Error details: {e}")
            print("Please ensure the model folder is complete and all dependencies are installed.")
            sys.exit(1)

    # ------------------------------------------------------
    # Text Preprocessing & Tokenization (Adapted for Mandarin)
    # ------------------------------------------------------
    def preprocess_text(self, text, mode='comparison', text_type='call'):
        """
        统一文本预处理函数，支持多种模式（普通话版本）
        
        Args:
            text: 输入文本
            mode: 处理模式
                - 'comparison': 用于比较分析（去除英文，script和call统一处理）
                - 'display': 用于显示（保留英文）
            text_type: 文本类型 ('call' 或 'script')，主要用于缓存区分
                
        Returns:
            处理后的文本
        """
        text_str = str(text)
        # 创建缓存键，包含模式和文本类型
        cache_key = f"{mode}_{text_type}_{text_str}"
        
        # 检查缓存
        if cache_key in self._preprocessed_texts_cache:
            return self._preprocessed_texts_cache[cache_key]
        
        if mode == 'comparison':
            # 比较模式：去除英文字母（script和call统一处理，确保比较一致性）
            processed = re.sub(r'[^\u4e00-\u9fff,.。，。%()（）]+', '', text_str)
        elif mode == 'display':
            # 显示模式：保留英文字母（用于sentence level analysis等显示用途）
            processed = re.sub(r'[^\u4e00-\u9fffa-zA-Z,.。，。%()（）]+', '', text_str)
        else:
            # 默认使用比较模式
            processed = re.sub(r'[^\u4e00-\u9fff,.。，。%()（）]+', '', text_str)
        
        # 标准化常见的语音转文字错误
        for correct, variations in self.error_patterns.items():
            for variation in variations:
                processed = processed.replace(variation, correct)
        
        # 普通话特定的标准化
        processed = re.sub(r'[,，]+', '，', processed)  # 标准化逗号
        processed = re.sub(r'[.。]+', '。', processed)  # 标准化句号
        
        result = processed.strip()
        
        # 缓存结果（限制缓存大小以防内存溢出）
        if len(self._preprocessed_texts_cache) < 10000:
            self._preprocessed_texts_cache[cache_key] = result
        
        return result

    def _build_reverse_synonym_map(self):
        """
        Build reverse synonym mapping for O(1) lookup performance.
        Maps each synonym to its canonical key and all related synonyms.
        """
        if self._reverse_synonym_map is not None:
            return self._reverse_synonym_map
            
        reverse_map = {}
        for key, synonyms in self.mandarin_synonyms.items():
            # Map the key to itself and all synonyms
            all_variants = {key} | set(synonyms)
            reverse_map[key] = all_variants
            
            # Map each synonym to the same set
            for synonym in synonyms:
                reverse_map[synonym] = all_variants
                
        self._reverse_synonym_map = reverse_map
        return reverse_map
    
    def expand_keywords(self, text):
        """
        Expand text with Mandarin synonyms for robust matching (optimized).
        - Tokenizes text using jieba.
        - Uses pre-computed reverse mapping for O(1) synonym lookup.
        """
        reverse_map = self._build_reverse_synonym_map()
        expanded_tokens = set()
        tokens = jieba.lcut(text)  # Using jieba for Mandarin
        
        for token in tokens:
            expanded_tokens.add(token)
            # Use reverse mapping for fast lookup
            if token in reverse_map:
                expanded_tokens.update(reverse_map[token])
        
        return expanded_tokens
    
    def tokenize_text(self, text):
        """
        Tokenize text using jieba and filter out stopwords.
        """
        tokens = jieba.lcut(text)  # Using jieba for Mandarin
        # Filter out stopwords and very short tokens
        business_tokens = [token for token in tokens if token not in self.stopwords and len(token) > 1]
        return business_tokens

    # ------------------------------------------------------
    # Multi-Product Weight Management
    # ------------------------------------------------------
    def detect_language_from_filename(self, file_path):
        """
        Detect language from CSV filename (Mandarin version only supports _M and _E)
        Examples:
        - "xxxxxx_M.wav.csv" -> "MAN" (Mandarin)
        - "xxxxxx_E.wav.csv" -> "ENG" (English)
        - "xxxxxx_C.wav.csv" -> Rejected (Cantonese not supported in Mandarin version)
        """
        filename = os.path.basename(file_path)
        # Remove .csv extension
        name_without_csv = filename.replace('.csv', '')
        # Remove .wav if present
        name_without_wav = name_without_csv.replace('.wav', '')
        
        # Get the last character after underscore
        if name_without_wav.endswith('_M'):
            return "MAN"
        elif name_without_wav.endswith('_E'):
            return "ENG"
        elif name_without_wav.endswith('_C'):
            print(f"❌ Error: Cantonese files (_C) are not supported in the Mandarin version.")
            print(f"Please use the Cantonese version (improved_call_coverage_checker.py) for _C files.")
            return None
        else:
            return None

    def detect_product_type_from_script(self, script_df, script_sheet_name=None):
        """
        Detect product type from script DataFrame based on script_sheet_name.
        Extracts product name from script_sheet_name (before underscore)
        Example: 'SID CPI3_MAN' -> product = 'SID CPI3', language = 'MAN'
        """
        if script_sheet_name and '_' in script_sheet_name:
            # Extract product name (before underscore)
            product_name = script_sheet_name.split('_')[0].strip()
            print(f"🔍 Detected product from sheet name '{script_sheet_name}': '{product_name}'")
            return product_name
        elif script_sheet_name:
            print(f"⚠️  Sheet name '{script_sheet_name}' doesn't contain underscore, using as-is: {script_sheet_name}")
            return script_sheet_name
        else:
            print(f"⚠️  No script_sheet_name provided, cannot detect product type")
            return "Unknown"

    def load_call_specific_weights(self, call_file_path, script_df, script_sheet_name=None):
        """
        Load specific weights and language-specific configurations based on call file and script
        """
        print("🔍 Loading call-specific weights and language configurations...")
        
        # Import shared dictionary
        import sys
        sys.path.insert(0, '..')
        import dictionaries as shared_dict
        
        # 1. Detect language from filename
        self.current_language = self.detect_language_from_filename(call_file_path)
        
        # 1b. If language detection from filename fails, try to detect from sheet name
        if not self.current_language and script_sheet_name:
            try:
                if '_' in script_sheet_name:
                    # Extract language code (after underscore)
                    language_code = script_sheet_name.split('_')[1].strip()
                    language_mapping = {'MAN': 'MAN', 'ENG': 'ENG'}  # Mandarin version only supports MAN and ENG
                    if language_code in language_mapping:
                        self.current_language = language_mapping[language_code]
                        print(f"🔍 Detected language from sheet name '{script_sheet_name}': '{self.current_language}'")
                    elif language_code == 'CAN':
                        print(f"❌ Error: Cantonese files (_CAN) are not supported in the Mandarin version.")
                        print(f"Please use the Cantonese version (improved_call_coverage_checker.py) for _CAN files.")
                        return None
            except Exception as e:
                print(f"⚠️  Error detecting language from sheet name: {e}")
        
        # 2. Load language-specific stopwords
        if self.current_language:
            try:
                self.stopwords = shared_dict.get_stopwords(self.current_language)
                print(f"✅ Loaded {self.current_language} stopwords: {len(self.stopwords)} words")
            except Exception as e:
                print(f"⚠️  Error loading {self.current_language} stopwords: {e}")
                print(f"⚠️  Using default Mandarin stopwords")
                self.stopwords = shared_dict.mandarin_stopwords
        else:
            # Default to Mandarin for this version
            self.stopwords = shared_dict.mandarin_stopwords
        
        # 3. Detect product type from script
        self.current_product = self.detect_product_type_from_script(script_df, script_sheet_name)
        
        # 4. Load corresponding weights
        # Trust get_product_weights to handle all fallback logic internally
        if self.current_language and self.current_product:
                self.current_weights = shared_dict.get_product_weights(
                    self.current_language, self.current_product
                )
                print(f"✅ Loaded weights for {self.current_language}:{self.current_product}")
                print(f"📊 Using {len(self.current_weights)} term weights")
        else:
            # Only fall back to general weights if language/product detection completely fails
            if not self.current_language:
                print(f"⚠️  Could not detect language from filename: {call_file_path}")
                if script_sheet_name:
                    print(f"⚠️  Also could not detect language from sheet name: {script_sheet_name}")
                else:
                    print(f"⚠️  No script sheet name provided for language detection")
            if not self.current_product:
                if script_sheet_name:
                    print(f"⚠️  Could not detect product type from script sheet name: {script_sheet_name}")
                else:
                    print(f"⚠️  No script sheet name provided for product detection")
            print(f"⚠️  Using general fallback weights")
            self.current_weights = getattr(shared_dict, 'term_importance', {})

    # ------------------------------------------------------
    # Salescall System Audio Recording Detection
    # ------------------------------------------------------
    def detect_system_audio_recordings(self, call_df):
        """
        Salescall Detection: Identify system audio recordings in sales calls.
        
        Logic:
        - Excludes speaker_1 and speaker_2 (main participants)
        - For remaining speakers, finds consecutive segments with:
          * 10+ consecutive rows by same speaker
          * Average sentence length >= 20 characters
        - If any segment meets criteria, marks ALL rows for that speaker as recordings
          (handles interruptions like coughing between recording segments)
        
        Args:
            call_df: DataFrame with call data
            
        Returns:
            pandas.Series: Boolean series indicating system audio recording rows
        """
        print("🔍 Salescall: Starting system audio recording detection...")
        
        speaker_col = 'Speaker Roles'
        text_col = 'Transcription'
        
        if speaker_col not in call_df.columns or text_col not in call_df.columns:
            print("⚠️  Salescall: Required columns not found, skipping recording detection")
            return pd.Series([False] * len(call_df), index=call_df.index)
        
        # Initialize recording detection array
        is_recording = pd.Series([False] * len(call_df), index=call_df.index)
        
        # Get all speakers except main participants (SPEAKER_1 and SPEAKER_2 are sales and customer)
        # Normalize speaker names to uppercase for consistent comparison
        all_speakers = call_df[speaker_col].dropna().unique()
        normalized_speakers = [str(s).upper() for s in all_speakers]
        main_speakers = ['SPEAKER_1', 'SPEAKER_2']
        system_speakers = [s for s in all_speakers if str(s).upper() not in main_speakers]
        
        print(f"📊 Salescall: Found {len(system_speakers)} potential system speakers: {system_speakers}")
        
        # Track speakers identified as system recordings
        recording_speakers = set()
        
        for speaker in system_speakers:
            speaker_rows = call_df[call_df[speaker_col] == speaker]
            if len(speaker_rows) < 10:  # Need at least 10 rows
                continue
            
            # Find consecutive segments for this speaker
            speaker_indices = speaker_rows.index.tolist()
            consecutive_segments = self._find_consecutive_segments(speaker_indices)
            
            # Check if any segment meets recording criteria
            speaker_is_recording = False
            for segment_indices in consecutive_segments:
                if len(segment_indices) >= 10:  # At least 10 consecutive rows
                    # Check average sentence length
                    segment_texts = call_df.loc[segment_indices, text_col].fillna('')
                    text_lengths = [len(str(text).strip()) for text in segment_texts]
                    avg_length = sum(text_lengths) / len(text_lengths) if text_lengths else 0
                    
                    if avg_length >= 20:  # Average length >= 20 characters
                        speaker_is_recording = True
                        print(f"🎵 Salescall: Detected recording segment - Speaker: {speaker}, "
                              f"Segment rows: {len(segment_indices)}, Avg length: {avg_length:.1f}")
                        break  # Found qualifying segment, no need to check others
            
            # If speaker qualifies, mark ALL their rows as recordings
            if speaker_is_recording:
                recording_speakers.add(speaker)
                all_speaker_indices = speaker_rows.index.tolist()
                is_recording.loc[all_speaker_indices] = True
                print(f"📢 Salescall: Marked ALL {len(all_speaker_indices)} rows for speaker '{speaker}' as recordings")
        
        total_recording_rows = is_recording.sum()
        print(f"✅ Salescall: Detection complete - {total_recording_rows} rows marked as system recordings")
        print(f"📊 Salescall: Recording speakers identified: {list(recording_speakers)}")
        
        return is_recording
    
    def _find_consecutive_segments(self, indices):
        """
        Salescall Helper: Find consecutive segments in a list of indices.
        
        Args:
            indices: List of row indices
            
        Returns:
            List of lists, each containing consecutive indices
        """
        if not indices:
            return []
        
        segments = []
        current_segment = [indices[0]]
        
        for i in range(1, len(indices)):
            if indices[i] == indices[i-1] + 1:  # Consecutive
                current_segment.append(indices[i])
            else:
                segments.append(current_segment)
                current_segment = [indices[i]]
        
        segments.append(current_segment)  # Add last segment
        return segments

    # ------------------------------------------------------
    # Unified Sentence Grouping Logic (3-Pass Strategy)
    # ------------------------------------------------------
    def build_grouped_lines(self, call_df, call_type="Sales Call"):
        """
        Unified grouped lines builder with configurable call type support.
        
        - call_type="Sales Call": Detects and excludes system audio recordings for grouping
        - call_type="SQCCB": Processes all rows without system recording detection
        - Uses three-pass physical merge strategy for optimal context grouping
        - Pass A: Forward merge for non-punctuated sentences
        - Pass B: Backward merge for short utterances  
        - Pass C: Forward merge for contextual windows
        """
        speaker_col, text_col = 'Speaker Roles', 'Transcription'
        start_col, end_col = 'Segment Start Time', 'Segment End Time'
        
        if speaker_col not in call_df.columns or text_col not in call_df.columns:
            raise KeyError(f"Required columns '{speaker_col}' and '{text_col}' not found.")
        
        # Create working copy to avoid modifying original DataFrame
        updated_call_df = call_df.copy()
        
        # Conditional system recording detection based on call_type
        if call_type == "Sales Call":
            # Sales Call mode: Detect and exclude system recordings
            system_recording_flags = self.detect_system_audio_recordings(call_df)
            updated_call_df['System_Audio_Recording'] = system_recording_flags.map({True: 'YES', False: 'NO'})
            
            # Filter out system recordings for grouping analysis
            print(f"📊 Sales Call: Filtering out {system_recording_flags.sum()} system recording rows from analysis")
            analysis_df = updated_call_df[~system_recording_flags].copy()
            print(f"📊 Sales Call: {len(analysis_df)} rows remaining for coverage analysis")
        else:
            # SQCCB mode: Skip system recording detection, process all rows
            print(f"🔍 SQCCB mode: Skip system recording detection, processing all {len(call_df)} rows")
            system_recording_flags = pd.Series([False] * len(call_df), index=call_df.index)
            updated_call_df['System_Audio_Recording'] = 'NO'
            analysis_df = updated_call_df.copy()

        def ends_with_punct(s):
            return s and s[-1] in '。！？；，.!?;'
        
        initial_rows = []
        for idx, row in analysis_df.iterrows():  # Use filtered DataFrame
            text = str(row.get(text_col, '')).strip()
            if not text: continue
            initial_rows.append({
                'orig_indices': [idx], 'text': text, 'speaker': row.get(speaker_col),
                'start_time': row.get(start_col), 'end_time': row.get(end_col),
            })

        if not initial_rows:
            print(f"⚠️  {call_type}: No valid rows found for grouping")
            return [], {}, updated_call_df

        # --- Pass A: Forward merge (non-punctuated) ---
        pass_a_results = []
        if initial_rows:
            pass_a_results.append(initial_rows[0].copy())
            for i in range(1, len(initial_rows)):
                current_row, last_group = initial_rows[i], pass_a_results[-1]
                if (last_group['speaker'] == current_row['speaker'] and 
                    not ends_with_punct(last_group['text']) and
                    (len(last_group['text']) + len(current_row['text']) + 1) <= 150):
                    last_group['text'] += ' ' + current_row['text']
                    last_group['orig_indices'].extend(current_row['orig_indices'])
                    if pd.notna(current_row['end_time']): last_group['end_time'] = current_row['end_time']
                else:
                    pass_a_results.append(current_row.copy())

        # --- Pass B: Backward merge (short utterances) ---
        pass_b_results = []
        if pass_a_results:
            for row in reversed(pass_a_results):
                if (pass_b_results and
                    pass_b_results[0]['speaker'] == row['speaker'] and
                    len(pass_b_results[0]['text']) < 20 and
                    (len(row['text']) + len(pass_b_results[0]['text']) + 1) <= 150):
                    first_group = pass_b_results[0]
                    first_group['text'] = row['text'] + ' ' + first_group['text']
                    first_group['orig_indices'] = row['orig_indices'] + first_group['orig_indices']
                    if pd.notna(row['start_time']): first_group['start_time'] = row['start_time']
                else:
                    pass_b_results.insert(0, row.copy())

        # --- Pass C: Forward contextual window merge ---
        final_results = []
        if pass_b_results:
            context_window_limit = 150  # Keep consistent with other limits
            final_results.append(pass_b_results[0].copy())
            for i in range(1, len(pass_b_results)):
                current_row, last_group = pass_b_results[i], final_results[-1]
                if (last_group['speaker'] == current_row['speaker'] and
                    (len(last_group['text']) + len(current_row['text']) + 1) <= context_window_limit):
                    last_group['text'] += ' ' + current_row['text']
                    last_group['orig_indices'].extend(current_row['orig_indices'])
                    if pd.notna(current_row['end_time']): last_group['end_time'] = current_row['end_time']
                else:
                    final_results.append(current_row.copy())

        # --- Final formatting ---
        grouped_lines, original_to_group = [], {}
        for group_id_counter, group in enumerate(final_results, 1):
            start_idx, end_idx = min(group['orig_indices']), max(group['orig_indices'])
            grouped_lines.append({
                'group_id': group_id_counter, 'text': group['text'], 'speaker': group['speaker'],
                'start_idx': start_idx, 'end_idx': end_idx,
                'start_time': group['start_time'], 'end_time': group['end_time'],
                'orig_indices': group['orig_indices']  # ✅ 添加orig_indices字段
            })
            for orig_idx in group['orig_indices']:
                original_to_group[orig_idx] = group_id_counter
            
        print(f"🔀 {call_type} grouping completed: {len(grouped_lines)} final groups created")
        
        # Return simplified results: grouped_lines, original_to_group, updated_call_df
        return grouped_lines, original_to_group, updated_call_df

    # ------------------------------------------------------
    # Legacy Sentence Grouping Logic (3-Pass Strategy)
    # ------------------------------------------------------
    def build_grouped_lines_for_analysis(self, analysis_df):
        """
        Build grouped lines for PURE human dialogue analysis (no system recordings).
        This function implements the three-pass physical merge strategy ONLY on human dialogue.
        
        Args:
            analysis_df: DataFrame containing ONLY human dialogue (no system recordings)
        
        Returns:
            grouped_lines: List of grouped human dialogue segments
            original_to_group: Mapping from original row indices to group IDs
        """
        speaker_col, text_col = 'Speaker Roles', 'Transcription'
        start_col, end_col = 'Segment Start Time', 'Segment End Time'
        
        if speaker_col not in analysis_df.columns or text_col not in analysis_df.columns:
            raise KeyError(f"Required columns '{speaker_col}' and '{text_col}' not found.")
        
        def ends_with_punct(s):
            return s and s[-1] in '。！？；，.!?;'
        
        initial_rows = []
        for idx, row in analysis_df.iterrows():  # Use pure human dialogue DataFrame
            text = str(row.get(text_col, '')).strip()
            if not text: continue
            initial_rows.append({
                'orig_indices': [idx], 'text': text, 'speaker': row.get(speaker_col),
                'start_time': row.get(start_col), 'end_time': row.get(end_col),
            })
        
        if not initial_rows:
            print(f"⚠️  No valid rows found for grouping in human dialogue")
            return [], {}
        
        # --- Pass A: Forward merge (non-punctuated) ---
        pass_a_results = []
        if initial_rows:
            pass_a_results.append(initial_rows[0].copy())
            for i in range(1, len(initial_rows)):
                current_row, last_group = initial_rows[i], pass_a_results[-1]
                if (last_group['speaker'] == current_row['speaker'] and
                    not ends_with_punct(last_group['text']) and
                    (len(last_group['text']) + len(current_row['text']) + 1) <= 150):
                    last_group['text'] += ' ' + current_row['text']
                    last_group['orig_indices'].extend(current_row['orig_indices'])
                    if pd.notna(current_row['end_time']): last_group['end_time'] = current_row['end_time']
                else:
                    pass_a_results.append(current_row.copy())
        
        # --- Pass B: Backward merge (short utterances) ---
        pass_b_results = []
        if pass_a_results:
            for row in reversed(pass_a_results):
                if (pass_b_results and 
                    pass_b_results[0]['speaker'] == row['speaker'] and
                    len(pass_b_results[0]['text']) < 20 and
                    (len(row['text']) + len(pass_b_results[0]['text']) + 1) <= 150):
                    first_group = pass_b_results[0]
                    first_group['text'] = row['text'] + ' ' + first_group['text']
                    first_group['orig_indices'] = row['orig_indices'] + first_group['orig_indices']
                    if pd.notna(row['start_time']): first_group['start_time'] = row['start_time']
                else:
                    pass_b_results.insert(0, row.copy())
        
        # --- Pass C: Forward contextual window merge ---
        final_results = []
        if pass_b_results:
            context_window_limit = 150  # Keep consistent with other limits
            final_results.append(pass_b_results[0].copy())
            for i in range(1, len(pass_b_results)):
                current_row, last_group = pass_b_results[i], final_results[-1]
                if (last_group['speaker'] == current_row['speaker'] and
                    (len(last_group['text']) + len(current_row['text']) + 1) <= context_window_limit):
                    last_group['text'] += ' ' + current_row['text']
                    last_group['orig_indices'].extend(current_row['orig_indices'])
                    if pd.notna(current_row['end_time']): last_group['end_time'] = current_row['end_time']
                else:
                    final_results.append(current_row.copy())
        
        # --- Final formatting ---
        grouped_lines, original_to_group = [], {}
        for group_id_counter, group in enumerate(final_results, 1):
            start_idx, end_idx = min(group['orig_indices']), max(group['orig_indices'])
            grouped_lines.append({
                'group_id': group_id_counter, 'text': group['text'], 'speaker': group['speaker'],
                'start_idx': start_idx, 'end_idx': end_idx,
                'start_time': group['start_time'], 'end_time': group['end_time'],
                'orig_indices': group['orig_indices']  # ✅ Include orig_indices
            })
            for orig_idx in group['orig_indices']:
                original_to_group[orig_idx] = group_id_counter
        
        print(f"🔀 Human dialogue grouping completed: {len(grouped_lines)} final groups created from {len(analysis_df)} human dialogue rows")
        
        # Return only grouped_lines and original_to_group (no updated_call_df)
        return grouped_lines, original_to_group
    

    

    
    # ------------------------------------------------------
    # Helper & Pattern Recognition Functions
    # ------------------------------------------------------
    def get_business_overlapping_keywords(self, text1, text2):
        """
        Find overlapping business keywords between two texts.
        Prioritizes important keywords first.
        """
        tokens1 = set(self.tokenize_text(text1))
        tokens2 = set(self.tokenize_text(text2))
        all_overlap = tokens1 & tokens2
        
        # Separate important and regular overlapping keywords
        important_overlap = all_overlap & self.important_keywords
        regular_overlap = all_overlap - important_overlap
        
        # Return important keywords first, then regular ones
        return list(important_overlap) + list(regular_overlap)

    def calculate_keyword_coverage(self, script_text, group_text):
        """
        Calculate keyword coverage ratio for business-critical terms.
        """
        script_tokens = set(self.tokenize_text(self.preprocess_text(script_text)))
        group_tokens = set(self.tokenize_text(self.preprocess_text(group_text)))
        
        # Focus on important keywords from script
        important_script_keywords = script_tokens & self.important_keywords
        if not important_script_keywords:
            return 0.0
        
        # Calculate coverage of important keywords
        covered_keywords = group_tokens & important_script_keywords
        return float(len(covered_keywords)) / len(important_script_keywords)

    def setup_enhanced_scoring_config(self):
        """
        Load enhanced scoring configuration from USER_ENHANCED_DISCUSSION_POINTS defined at top of file.
        
        Returns the configuration dictionary for pattern-based score enhancements.
        
        Usage: Configure discussion points at the top of this file by modifying the
        USER_ENHANCED_DISCUSSION_POINTS dictionary.
        """
        return USER_ENHANCED_DISCUSSION_POINTS
    
    def detect_date_patterns(self, text):
        """
        Detect various Chinese date patterns in text including mixed Arabic/Chinese numerals.
        
        Supports formats like:
        - 2025年10月9日
        - 二零二五年十月九日  
        - 10月9日
        - 十月九日
        - 2025年十月九日 (mixed)
        - 二零二五年10月9日 (mixed)
        - 2025-10-09, 2025/10/09
        """
        # Quick check for date indicators
        if not any(indicator in text for indicator in ['年', '月', '日', '/', '-']):
            return False, []
        
        # Comprehensive date patterns for Chinese dates
        patterns = [
            # Full Chinese dates with years, months, days
            r'[二零一三四五六七八九]{2,4}年[一二三四五六七八九十\d]{1,3}月[一二三四五六七八九十\d]{1,3}日',
            # Mixed Arabic year with Chinese month/day
            r'\d{4}年[一二三四五六七八九十\d]{1,3}月[一二三四五六七八九十\d]{1,3}日',
            # Chinese year with Arabic month/day  
            r'[二零一三四五六七八九]{2,4}年\d{1,2}月\d{1,2}日',
            # Standard Arabic format
            r'\d{4}年\d{1,2}月\d{1,2}日',
            # Month-day only formats
            r'[一二三四五六七八九十\d]{1,3}月[一二三四五六七八九十\d]{1,3}日',
            r'\d{1,2}月\d{1,2}日',
            # Western date formats
            r'\d{4}[-/]\d{1,2}[-/]\d{1,2}'
        ]
        
        found_dates = []
        for pattern in patterns:
            matches = re.findall(pattern, text)
            found_dates.extend(matches)
        
        return len(found_dates) > 0, found_dates[:3]  # Return first 3 matches
    
    def detect_numeric_patterns(self, text):
        """
        Detect financial/numeric patterns while avoiding date-like sequences.
        
        Looks for:
        - Monetary amounts (元, 万, 千, 亿)
        - Currency symbols (¥, $, €)
        - Percentages
        - Large numbers
        """
        # Quick check for numeric indicators
        if not re.search(r'[\d¥$€%万千百亿元]', text):
            return False, []
        
        patterns = [
            # Chinese monetary amounts  
            r'[港人民币港币]{1,3}\s*\d+(?:\.\d+)?[万千百亿元]?',
            # Currency symbols with amounts
            r'[¥$€]\s*[\d,]+(?:\.\d+)?',
            # Standalone amounts with Chinese units
            r'\d+(?:\.\d+)?[万千百亿元]',
            # Percentages
            r'\d+(?:\.\d+)?%'
        ]
        
        found_numbers = []
        for pattern in patterns:
            matches = re.findall(pattern, text)
            found_numbers.extend(matches)
        
        return len(found_numbers) > 0, found_numbers[:2]  # Return first 2 matches
    
    def apply_pattern_enhancement(self, base_score, text, discussion_point, dynamic_numeric_boost=0):
        """
        Apply pattern-based score enhancement for specific discussion points.
        This is called automatically during similarity calculation.
        
        How it works:
        1. Checks if the discussion_point is configured for enhancement
        2. If configured, detects date/numeric patterns in the text
        3. Applies the configured boost weights if patterns are found
        4. If no static config, applies dynamic numeric boost if available
        5. Returns the enhanced score (capped at 1.0)
        
        Configuration: Modify USER_ENHANCED_DISCUSSION_POINTS at the top of this file
        to specify which discussion points should receive which type of boosts.
        
        Weight Guidelines:
        - date_boost: 0.05-0.25 for discussion points about dates/deadlines
        - numeric_boost: 0.05-0.25 for discussion points about prices/amounts
        - dynamic_numeric_boost: 0.15 for auto-detected price/interest related points (English keywords only)
        
        Efficiency: The system efficiently skips pattern detection for discussion
        points that are not configured for enhancement, saving computation time.
        """
        enhancement_config = self.setup_enhanced_scoring_config()
        enhanced_score = base_score
        
        # Priority 1: Use static configuration if available
        if discussion_point in enhancement_config:
            config = enhancement_config[discussion_point]

            # Apply date pattern boost if configured
            if config.get('date_boost', 0) > 0:
                has_dates, _ = self.detect_date_patterns(text)
                if has_dates:
                    enhanced_score += config['date_boost']

            # Apply numeric pattern boost if configured  
            if config.get('numeric_boost', 0) > 0:
                has_numbers, _ = self.detect_numeric_patterns(text)
                if has_numbers:
                    enhanced_score += config['numeric_boost']
        
        # Priority 2: Use dynamic numeric boost if no static config and dynamic boost is available
        elif dynamic_numeric_boost > 0:
            has_numbers, _ = self.detect_numeric_patterns(text)
            if has_numbers:
                enhanced_score += dynamic_numeric_boost

        return min(1.0, enhanced_score)  # Cap at 1.0

    def calculate_token_rouge_l(self, text1, text2):
        """
        Calculate ROUGE-L (Longest Common Subsequence) F-score at character level.
        """
        seq1 = list(text1)
        seq2 = list(text2)
        
        if not seq1 or not seq2:
            return 0.0
        
        n, m = len(seq1), len(seq2)
        dp = [[0] * (m + 1) for _ in range(n + 1)]
        
        # Dynamic programming to find LCS length
        for i in range(1, n + 1):
            for j in range(1, m + 1):
                if seq1[i-1] == seq2[j-1]:
                    dp[i][j] = dp[i-1][j-1] + 1
                else:
                    dp[i][j] = max(dp[i-1][j], dp[i][j-1])
        
        lcs_len = dp[n][m]
        if lcs_len == 0:
            return 0.0
        
        # Calculate precision, recall, and F-score
        precision = lcs_len / m
        recall = lcs_len / n
        
        if precision + recall == 0:
            return 0.0
        
        return (2 * precision * recall) / (precision + recall)

    # ------------------------------------------------------
    # Core Semantic Similarity Calculation (UPGRADED TO SBERT)
    # ------------------------------------------------------
    def calculate_semantic_similarity(self, text1, text2):
        """
        Calculates semantic similarity using a hybrid system led by SBERT.
        
        Metrics Breakdown:
        1. SBERT Semantic Score (60%): Deep semantic understanding using transformer embeddings
        2. Expanded Token Overlap (15%): Synonym-aware matching for business terms
        3. ROUGE-L (15%): Sequence similarity (LCS-based character matching)
        4. Keyword Coverage (10%): Compliance check for critical business terms
        
        Returns:
            dict: Dictionary containing all individual metrics and the final weighted score
        """
        # Use cache to avoid re-computing for the same text pair
        cache_key = (text1, text2)
        if cache_key in self._similarity_cache:
            return self._similarity_cache[cache_key]

        clean1 = self.preprocess_text(text1)
        clean2 = self.preprocess_text(text2)
        metrics = {}
        
        # 1. SBERT Semantic Score (The new core engine)
        try:
            embeddings = self.sbert_model.encode([clean1, clean2])
            semantic_score = util.cos_sim(embeddings[0], embeddings[1])[0][0].item()
        except Exception as e:
            print(f"⚠️  SBERT encoding failed for texts: '{clean1[:20]}...' | '{clean2[:20]}...'. Error: {e}. Falling back to 0.0")
            semantic_score = 0.0
        metrics['semantic_score'] = semantic_score
        
        # 2. Expanded token overlap (with synonyms) - Still valuable for business logic
        expanded1 = self.expand_keywords(clean1)
        expanded2 = self.expand_keywords(clean2)
        union_size = len(expanded1 | expanded2)
        expanded_similarity = len(expanded1 & expanded2) / union_size if union_size > 0 else 0
        metrics['expanded_overlap'] = expanded_similarity
        
        # 3. Token-level ROUGE-L (LCS-based sequence similarity)
        rouge_l = self.calculate_token_rouge_l(clean1, clean2)
        metrics['rouge_l'] = rouge_l

        # 4. Keyword coverage (bounded 0..1)
        keyword_coverage = self.calculate_keyword_coverage(text1, text2)
        metrics['keyword_coverage'] = keyword_coverage
        
        # New weighted combination with SBERT as the primary component
        weighted_score = (
            semantic_score * 0.60 +          # SBERT provides the main semantic signal
            expanded_similarity * 0.15 +     # Synonyms provide explicit business logic
            rouge_l * 0.15 +                 # Sequence matching guards against nonsensical order
            keyword_coverage * 0.10          # Critical keywords act as a compliance backstop
        )
        metrics['weighted_score'] = weighted_score
        
        # Update cache (limit size to prevent memory bloat)
        if len(self._similarity_cache) < 5000:
            self._similarity_cache[cache_key] = metrics

        return metrics
    
    # ------------------------------------------------------
    # Script & Coverage Logic (Adapted for SBERT)
    # ------------------------------------------------------
    def parse_script_variations(self, script_text):
        """
        Parse script variations from text, handling multiple formats and versions.
        
        Args:
            script_text (str): Raw script text that may contain multiple variations
            
        Returns:
            list: List of individual script variations
        """
        if pd.isna(script_text):
            return []
        
        script_text = str(script_text).strip()
        if not script_text:
            return []
        
        # Split by common separators and version indicators
        raw_splits = re.split(r'[.;\n]|版本[AB][:：]|Version [AB][:：]|情况[一二三四五六七八九十][:：]', script_text)
        
        variations = []
        for split in raw_splits:
            split = split.strip()
            if split and len(split) > 5:  # Filter out very short fragments
                if len(split) > 50:  # Further split long segments
                    sub_splits = [s.strip() for s in split.split(',') if s.strip() and len(s.strip()) > 5]
                    variations.extend(sub_splits)
                else:
                    variations.append(split)
        
        # Remove duplicates while preserving order
        unique_variations = []
        seen = set()
        for var in variations:
            if var not in seen:
                unique_variations.append(var)
                seen.add(var)
        
        return unique_variations if unique_variations else [script_text]

    def check_coverage(self, call_df, required_points_df, grouped_lines, threshold=0.4):
        """
        Checks coverage using the SBERT-powered similarity engine.
        
        Args:
            call_df: DataFrame with call transcription data
            required_points_df: DataFrame with required discussion points and scripts
            grouped_lines: List of grouped call text segments
            threshold: Minimum similarity score to consider a point "covered"
            
        Returns:
            pd.DataFrame: Coverage analysis results
        """
        results = []
        
        # Group scripts by discussion point
        point_to_scripts = {}
        for _, row in required_points_df.iterrows():
            point = row['Required_Discussion_Point']
            script = row['Standard_Script']
            if pd.notna(point) and pd.notna(script):
                point_to_scripts.setdefault(point, []).append(str(script))

        # Analyze coverage for each discussion point
        for point, scripts in point_to_scripts.items():
            result = self._analyze_single_point_coverage(point, scripts, grouped_lines, threshold)
            results.append(result)
        
        return pd.DataFrame(results)
    
    def _analyze_single_point_coverage(self, point, scripts, grouped_lines, threshold):
        """
        SBERT优化版本：采用双重验证策略
        
        IMPLEMENTS DUAL VERIFICATION: Holistic vs Granular matching strategy.
        
        策略：
        1. 解析脚本变体得到"精准小靶子"（Granular matching）
        2. 创建完整脚本文本用于整段匹配（Holistic matching）
        3. 对每个对话组进行双重匹配，取最高分
        4. 追踪匹配类型和分数用于诊断
        
        Args:
            point: Discussion point name
            scripts: List of standard scripts for this point
            grouped_lines: List of grouped call text segments (已包含丰富上下文)
            threshold: Coverage threshold
            
        Returns:
            dict: Analysis results for this discussion point
        """
        # 步骤1: 解析脚本变体，得到"精准小靶子"列表（用于Granular matching）
        script_variations = []
        for script in scripts:
            script_variations.extend(self.parse_script_variations(script))
        
        # 步骤2: 创建完整脚本文本（用于Holistic matching）
        holistic_script_text = ' '.join(scripts) if scripts else ""
        
        # Dynamic numeric boost detection
        dynamic_numeric_boost = 0
        
        # Dynamic rule 1: Check point name for price/floating rate keywords
        point_lower = point.lower()
        if 'price' in point_lower or 'floating rate' in point_lower:
            dynamic_numeric_boost = 0.15
        
        # Dynamic rule 2: Check script content for price/interest/% keywords
        if dynamic_numeric_boost == 0 and scripts:
            combined_scripts_lower = ' '.join(scripts).lower()
            if any(keyword in combined_scripts_lower for keyword in ['price', 'interest', '%']):
                dynamic_numeric_boost = 0.15
        
        # 初始化最佳匹配结果
        best_score = 0.0
        best_original_score = 0.0
        best_group_info = {}
        best_variation = ""
        best_metrics = {}
        best_match_type = ""  # Track whether final match was Holistic or Granular
        best_holistic_score = 0  # Track holistic score for the best match
        best_granular_score = 0  # Track granular score for the best match
        
        # Get point-specific weights for this discussion point
        try:
            import sys
            sys.path.insert(0, '..')
            import dictionaries as shared_dict
            point_weights = shared_dict.get_point_specific_weights(
                self.current_language, self.current_product, point
            )
            # Temporarily switch to point-specific weights
            original_weights = self.current_weights
            self.current_weights = point_weights
            point_specific_mode = True
        except Exception as e:
            # Fallback to existing weights if point-specific weights fail
            point_specific_mode = False
        
        # 步骤3: 遍历包含丰富上下文的对话组，进行双重验证
        for group_info in grouped_lines:
            group_text = group_info['text']
            if not group_text:
                continue
            
            # Step 1: Holistic matching - compare with complete script text
            holistic_metrics = self.calculate_semantic_similarity(holistic_script_text, group_text)
            holistic_weighted_score = holistic_metrics['weighted_score']
            
            # Step 2: Granular matching - find best among script variations
            best_granular_score_for_group = 0
            best_granular_metrics = {}
            best_granular_variation = ""
            
            for variation in script_variations:
                metrics = self.calculate_semantic_similarity(variation, group_text)
                granular_score = metrics['weighted_score']
                
                if granular_score > best_granular_score_for_group:
                    best_granular_score_for_group = granular_score
                    best_granular_metrics = metrics
                    best_granular_variation = variation
            
            # Step 3: Choose the better approach for this group
            if holistic_weighted_score > best_granular_score_for_group:
                final_score_for_group = holistic_weighted_score
                final_metrics_for_group = holistic_metrics
                match_type_for_group = "Holistic"
                final_variation_for_group = holistic_script_text
            else:
                final_score_for_group = best_granular_score_for_group
                final_metrics_for_group = best_granular_metrics
                match_type_for_group = "Granular"
                final_variation_for_group = best_granular_variation
            
            # Apply pattern enhancement
            enhanced_score = self.apply_pattern_enhancement(final_score_for_group, group_text, point, dynamic_numeric_boost)
            
            # Update best match if this group scored higher
            if enhanced_score > best_score:
                best_score = enhanced_score
                best_original_score = final_score_for_group
                best_group_info = group_info
                best_metrics = final_metrics_for_group
                best_variation = final_variation_for_group
                best_match_type = match_type_for_group
                best_holistic_score = holistic_weighted_score
                best_granular_score = best_granular_score_for_group
                
                # Early exit for very high scores (SBERT can achieve higher scores)
                if best_score >= 0.95:
                    break
            
            if best_score >= 0.95:
                break
        
        # Restore original weights if we switched to point-specific mode
        if point_specific_mode:
            self.current_weights = original_weights
        
        # Get overlapping keywords for the best match
        overlapping_keywords = self.get_business_overlapping_keywords(
            self.preprocess_text(best_variation or (scripts[0] if scripts else "")),
            best_group_info.get('text', '')
        )
        
        enhancement_boost = best_score - best_original_score
        
        return {
            'Required_Discussion_Point': point,
            'Covered': 'Covered' if best_score >= threshold else 'Not Covered',
            'Weighted_Score': round(best_score, 3),
            'Original_Score': round(best_original_score, 3),
            'Enhancement_Boost': round(enhancement_boost, 3),
            'SBERT_Semantic_Score': round(best_metrics.get('semantic_score', 0), 3),  # New metric
            'Expanded_Overlap': round(best_metrics.get('expanded_overlap', 0), 3),
            'ROUGE_L': round(best_metrics.get('rouge_l', 0), 3),
            'Keyword_Coverage': round(best_metrics.get('keyword_coverage', 0), 3),
            'Overlapping_Keywords': ', '.join(overlapping_keywords),
            'Matched_Group': best_group_info.get('text', ''),
            'Group_ID': best_group_info.get('group_id', -1),
            'Speaker': best_group_info.get('speaker', ''),
            'Best_Matching_Variation': best_variation or (scripts[0] if scripts else ''),
            'All_Variations_Count': len(script_variations),
            'Match_Type': best_match_type,  # New: Holistic or Granular
            'Holistic_Score': round(best_holistic_score, 3),  # New: holistic matching score
            'Granular_Score': round(best_granular_score, 3),  # New: granular matching score
        }

    # ------------------------------------------------------
    # Reporting & Output Functions
    # ------------------------------------------------------
    def create_grouped_call_dataframe(self, grouped_lines):
        """Create DataFrame from grouped lines for export."""
        return pd.DataFrame(grouped_lines)
    
    def create_call_text_analysis_view(self, grouped_lines, script_df, threshold=0.4):
        """
        Create detailed call text analysis view showing how each group matches against all discussion points.
        
        Args:
            grouped_lines: List of grouped call text segments
            script_df: DataFrame with script data
            threshold: Threshold for determining hits
            
        Returns:
            pd.DataFrame: Analysis view with columns for each discussion point
        """
        analysis_rows = []
        
        # Group script variations by discussion point and create holistic scripts
        point_to_variations = {}
        point_to_holistic_script = {}
        for _, row in script_df.iterrows():
            point = row['Required_Discussion_Point']
            script = row['Standard_Script']
            if pd.notna(point) and pd.notna(script):
                if point not in point_to_variations:
                    point_to_variations[point] = []
                    point_to_holistic_script[point] = []
                point_to_variations[point].extend(self.parse_script_variations(str(script)))
                point_to_holistic_script[point].append(str(script))
        
        # Create complete holistic script text for each point
        point_to_complete_holistic = {}
        for point, scripts in point_to_holistic_script.items():
            point_to_complete_holistic[point] = ' '.join(scripts)
        
        all_points = list(point_to_variations.keys())

        # Analyze each group against all discussion points
        for group in grouped_lines:
            analysis_row = {
                'Group_ID': group['group_id'],
                'Speaker': group['speaker'],
                'Call_Text': group['text']
            }
            
            # Check each discussion point for this group
            for point in all_points:
                # Dynamic numeric boost detection for this point
                dynamic_numeric_boost = 0
                
                # Dynamic rule 1: Check point name for price/floating rate keywords
                point_lower = point.lower()
                if 'price' in point_lower or 'floating rate' in point_lower:
                    dynamic_numeric_boost = 0.15
                
                # Dynamic rule 2: Check script content for price/interest/% keywords
                if dynamic_numeric_boost == 0:
                    holistic_script = point_to_complete_holistic.get(point, "")
                    if holistic_script:
                        combined_scripts_lower = holistic_script.lower()
                        if any(keyword in combined_scripts_lower for keyword in ['price', 'interest', '%']):
                            dynamic_numeric_boost = 0.15
                
                # Dual verification: Holistic vs Granular matching
                script_variations = point_to_variations.get(point, [])
                holistic_script = point_to_complete_holistic.get(point, "")
                
                # Step 1: Holistic matching
                holistic_score = 0
                holistic_keywords = []
                if holistic_script:
                    holistic_metrics = self.calculate_semantic_similarity(holistic_script, group['text'])
                    holistic_score = self.apply_pattern_enhancement(
                        holistic_metrics['weighted_score'], group['text'], point, dynamic_numeric_boost
                    )
                    holistic_keywords = self.get_business_overlapping_keywords(
                        self.preprocess_text(holistic_script), group['text']
                    )
                
                # Step 2: Granular matching
                granular_score = 0
                granular_keywords = []
                for variation in script_variations:
                    metrics = self.calculate_semantic_similarity(variation, group['text'])
                    enhanced_score = self.apply_pattern_enhancement(
                        metrics['weighted_score'], group['text'], point, dynamic_numeric_boost
                    )
                    
                    if enhanced_score > granular_score:
                        granular_score = enhanced_score
                        granular_keywords = self.get_business_overlapping_keywords(
                            self.preprocess_text(variation), group['text']
                        )
                
                # Step 3: Choose the better approach
                if holistic_score > granular_score:
                    best_score_for_point = holistic_score
                    best_keywords_for_point = holistic_keywords
                else:
                    best_score_for_point = granular_score
                    best_keywords_for_point = granular_keywords
                
                # Determine if this is a hit
                point_hit = best_score_for_point >= threshold
                
                # Create shortened point name for column
                point_short = f"{point[:30]}{'...' if len(point) > 30 else ''}"
                
                # Add columns for this discussion point
                analysis_row[f"{point_short}_Hit"] = 'YES' if point_hit else 'NO'
                analysis_row[f"{point_short}_Score"] = round(best_score_for_point, 3)
                analysis_row[f"{point_short}_Keywords"] = ', '.join(best_keywords_for_point)
            
            analysis_rows.append(analysis_row)
        
        return pd.DataFrame(analysis_rows)

    def create_call_text_analysis_view_with_separation(self, grouped_lines, system_audio_df, script_df, threshold=0.4):
        """
        STEP 4: Late Merge & Reporting - Create call text analysis with proper separation handling.
        
        This function implements the late merge strategy:
        1. Generate analysis results for human dialogue groups (grouped_lines) - includes Hit, Score, Keywords
        2. Format system recording rows with minimal structure (_Hit='NO' only, no Score/Keywords)
        3. Combine both into a single comprehensive report
        """
        analysis_rows = []
        
        # Group script variations by discussion point
        point_to_variations = {}
        for _, row in script_df.iterrows():
            point = row['Required_Discussion_Point']
            script = row['Standard_Script']
            if pd.notna(point) and pd.notna(script):
                if point not in point_to_variations:
                    point_to_variations[point] = []
                point_to_variations[point].extend(self.parse_script_variations(str(script)))
        
        all_points = list(point_to_variations.keys())
        
        # Part 1: Process human dialogue groups (with actual analysis)
        for group in grouped_lines:
            analysis_row = {
                'Group_ID': group['group_id'], 
                'Speaker': group['speaker'], 
                'Call_Text': group['text'],
                'System_Audio_Recording': 'NO'
            }
            
            # Analyze each discussion point for this group
            for point in all_points:
                point_short = f"{point[:30]}{'...' if len(point) > 30 else ''}"
                script_variations = point_to_variations.get(point, [])
                
                best_score_for_point = 0
                best_keywords_for_point = []
                
                for variation in script_variations:
                    metrics = self.calculate_semantic_similarity(variation, group['text'])
                    granular_score = metrics['weighted_score']
                    
                    if granular_score > best_score_for_point:
                        best_score_for_point = granular_score
                        # Extract keywords using jieba for Mandarin
                        import jieba
                        words = list(jieba.cut(group['text']))
                        # Use proper stopwords from dictionaries based on language
                        import dictionaries
                        stopwords = dictionaries.get_stopwords('MAN')  # Mandarin version
                        keywords = [w for w in words if len(w) > 1 and w not in stopwords]
                        from collections import Counter
                        word_freq = Counter(keywords)
                        best_keywords_for_point = [word for word, count in word_freq.most_common(10)]
            
                # Record results for this point
                point_hit = best_score_for_point >= threshold
                analysis_row[f"{point_short}_Hit"] = 'YES' if point_hit else 'NO'
                analysis_row[f"{point_short}_Score"] = round(best_score_for_point, 3)
                analysis_row[f"{point_short}_Keywords"] = ', '.join(best_keywords_for_point)
            
            analysis_rows.append(analysis_row)
        
        # Part 2: Process system recording rows (only Hit column needed)
        if not system_audio_df.empty:
            for idx, row in system_audio_df.iterrows():
                system_row = {
                    'Group_ID': f"SYS_{idx}",
                    'Speaker': row.get('Speaker Roles', 'System'),
                    'Call_Text': str(row.get('Transcription', '')),
                    'System_Audio_Recording': 'YES'
                }
                
                # For system recordings, only Hit column is needed
                for point in all_points:
                    point_short = f"{point[:30]}{'...' if len(point) > 30 else ''}"
                    system_row[f"{point_short}_Hit"] = 'NO'
                
                analysis_rows.append(system_row)
        
        return pd.DataFrame(analysis_rows)

    def create_sentence_level_output(self, call_df, original_to_group):
        """
        Create sentence-level output showing original to group mapping.
        
        Args:
            call_df: Original call DataFrame
            original_to_group: Mapping from original row indices to group IDs
            
        Returns:
            pd.DataFrame: Sentence-level analysis
        """
        speaker_col = 'Speaker Roles'
        text_col = 'Transcription'
        start_col = 'Segment Start Time'
        end_col = 'Segment End Time'
        
        sentence_level_data = []
        
        for idx, row in call_df.iterrows():
            text_val = row.get(text_col)
            if pd.isna(text_val) or not str(text_val).strip():
                continue
                
            sentence_level_data.append({
                'Original_Row': idx,
                'Group_ID': original_to_group.get(idx),
                'Speaker': row.get(speaker_col, ''),
                'Start_Time': row.get(start_col),
                'End_Time': row.get(end_col),
                'Original_Text': str(text_val).strip()
            })
        
        return pd.DataFrame(sentence_level_data)

# Old main function removed - using new four-step strategy below


# ==========================================================
# Executive Summary Report Generation
# ==========================================================

def parse_time_to_seconds(time_str):
    """
    Parse time string in format "00:18:00" to total seconds.
    Extracts the last 6 characters (MM:SS) and converts to seconds.
    """
    if pd.isna(time_str) or not time_str:
        return 0
    
    try:
        time_str = str(time_str).strip()
        # Extract last 6 characters for MM:SS format
        if len(time_str) >= 6:
            time_part = time_str[-6:]  # Get "18:00" from "00:18:00"
        else:
            time_part = time_str
        
        # Split and convert to seconds
        if ':' in time_part:
            parts = time_part.split(':')
            if len(parts) == 2:
                minutes = int(parts[0])
                seconds = int(parts[1])
                return minutes * 60 + seconds
            elif len(parts) == 3:  # Handle full HH:MM:SS
                hours = int(parts[0])
                minutes = int(parts[1])
                seconds = int(parts[2])
                return hours * 3600 + minutes * 60 + seconds
        
        # If no colon, try to parse as plain number
        return float(time_str)
    except (ValueError, IndexError):
        print(f"⚠️ Warning: Could not parse time format: {time_str}")
        return 0

def generate_executive_summary_report(results_df, updated_call_df, script_df, grouped_lines, output_file=None):
    """
    Generate a consolidated Executive Summary Excel report with three sections:
    A: Overall Performance KPIs
    B: Key Insights - Speaker View  
    C: Detailed Coverage List
    
    Returns the final DataFrame that was written to Excel.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from collections import Counter
    import string
    
    print("📊 Generating Executive Summary report...")
    
    # ==========================================
    # Section A: Overall Performance KPIs
    # ==========================================
    print("📈 Calculating Overall Performance KPIs...")
    
    # Get total call duration directly from the last row's segment end time
    if 'Segment End Time' in updated_call_df.columns and not updated_call_df.empty:
        # Get the last row's end time value directly (keep as string if it's string)
        total_duration = updated_call_df['Segment End Time'].iloc[-1]
        if pd.isna(total_duration):
            total_duration = "N/A"
        else:
            total_duration = str(total_duration)
    else:
        total_duration = "N/A"
    
    # Calculate coverage metrics
    total_points = len(results_df)
    covered_points = len(results_df[results_df['Covered'] == 'Covered'])
    uncovered_points = total_points - covered_points
    coverage_rate = (covered_points / total_points * 100) if total_points > 0 else 0
    
    kpis_data = {
        'Metric': [
            'Total Call Duration',
            'Compliance Coverage Rate (%)',
            'Total Points Checked',
            'Covered Points',
            'Uncovered Risk Points'
        ],
        'Value': [
            str(total_duration),
            f"{coverage_rate:.1f}%",
            str(total_points),
            str(covered_points),
            str(uncovered_points)
        ]
    }
    kpis_df = pd.DataFrame(kpis_data)
    
    # ==========================================
    # Section B: Key Insights - Speaker View
    # ==========================================
    print("🎤 Generating Speaker View insights...")
    
    # Filter out system recordings for speaker analysis
    non_system_df = updated_call_df[updated_call_df.get('System_Audio_Recording', 'NO') == 'NO'].copy()
    
    # Calculate speaker statistics
    speaker_stats = []
    if not non_system_df.empty and 'Speaker Roles' in non_system_df.columns and 'Transcription' in non_system_df.columns:
        # Group by speaker and calculate metrics
        for speaker in non_system_df['Speaker Roles'].unique():
            if pd.isna(speaker):
                continue
                
            speaker_data = non_system_df[non_system_df['Speaker Roles'] == speaker]
            
            # Calculate speaking duration using proper time parsing with strptime
            if 'Segment Start Time' in speaker_data.columns and 'Segment End Time' in speaker_data.columns:
                speaking_duration = 0.0
                
                for _, segment in speaker_data.iterrows():
                    start_time_str = str(segment.get('Segment Start Time', '')).strip()
                    end_time_str = str(segment.get('Segment End Time', '')).strip()
                    
                    if start_time_str and end_time_str and start_time_str != 'nan' and end_time_str != 'nan':
                        try:
                            from datetime import datetime
                            # Parse time strings as HH:MM:SS format
                            start_time = datetime.strptime(start_time_str, "%H:%M:%S")
                            end_time = datetime.strptime(end_time_str, "%H:%M:%S")
                            
                            # Calculate duration in seconds
                            duration_seconds = (end_time - start_time).total_seconds()
                            if duration_seconds >= 0:  # Only add positive durations
                                speaking_duration += duration_seconds
                                
                        except ValueError:
                            # If parsing fails, try to use the old parse_time_to_seconds as fallback
                            try:
                                start_seconds = parse_time_to_seconds(start_time_str)
                                end_seconds = parse_time_to_seconds(end_time_str)
                                duration_seconds = end_seconds - start_seconds
                                if duration_seconds >= 0:
                                    speaking_duration += duration_seconds
                            except:
                                continue  # Skip this segment if can't parse
                
                print(f"🔍 Debug - Speaker: {speaker}, Speaking Duration: {speaking_duration:.1f}s")
            else:
                speaking_duration = 0.0
                print(f"🔍 Debug - Speaker: {speaker} - Time columns not found")
            
            # Calculate word count and extract text
            all_text = ' '.join(speaker_data['Transcription'].fillna('').astype(str))
            word_count = len(all_text.replace(' ', ''))  # Character count for Chinese text
            
            # Calculate words per second
            words_per_second = (word_count / speaking_duration) if speaking_duration > 0 else 0
            
            # Generate top 10 keywords using jieba tokenization for Mandarin
            import jieba
            # Tokenize using jieba and filter
            words = jieba.lcut(all_text)
            words = [w.strip() for w in words if len(w.strip()) >= 2]  # Keep words with 2+ characters
            
            # Use proper stopwords from dictionaries based on language
            import dictionaries
            stopwords = dictionaries.get_stopwords('MAN')  # Mandarin version
            filtered_words = [w for w in words if w not in stopwords]
            
            # Get top 10 most frequent words
            word_counter = Counter(filtered_words)
            top_10_words = [word for word, count in word_counter.most_common(10)]
            
            # Format as requested: "{'word1', 'word2', ..., 'word10'}"
            keywords_str = "{" + ", ".join([f"'{word}'" for word in top_10_words]) + "}"
            
            speaker_stats.append({
                'Speaker': speaker,
                'Speaking Duration (seconds)': speaking_duration,
                'Word Count': str(word_count),
                'Words per Second': f"{words_per_second:.2f}",
                'Top 10 Keywords': keywords_str
            })
    
    # Identify system recording speakers and add them to speaker_stats
    system_recording_speakers = set()
    if 'System_Audio_Recording' in updated_call_df.columns:
        system_df = updated_call_df[updated_call_df['System_Audio_Recording'] == 'YES']
        if not system_df.empty and 'Speaker Roles' in system_df.columns:
            system_recording_speakers = set(system_df['Speaker Roles'].dropna().unique())
            
            # Add system recording speakers to the speaker_stats
            for sys_speaker in system_recording_speakers:
                if sys_speaker and sys_speaker not in [stat['Speaker'] for stat in speaker_stats]:
                    speaker_stats.append({
                        'Role': 'System',
                        'Speaker': str(sys_speaker),
                        'Speaking Duration (seconds)': 'N/A',
                        'Word Count': 'N/A',
                        'Words per Second': 'N/A',
                        'Top 10 Keywords': 'N/A'
                    })
    
    # Sort speakers by word count to identify Sales and Customer (excluding system speakers)
    regular_speakers = [stat for stat in speaker_stats if stat.get('Role') != 'System']
    if regular_speakers:
        # Convert word count back to int for sorting
        for stat in regular_speakers:
            stat['_word_count_int'] = int(stat['Word Count'])
        
        regular_speakers.sort(key=lambda x: x['_word_count_int'], reverse=True)
        
        # Assign roles: highest word count = Sales, second highest = Customer
        if len(regular_speakers) >= 1:
            regular_speakers[0]['Role'] = 'Sales'
        if len(regular_speakers) >= 2:
            regular_speakers[1]['Role'] = 'Customer'
        
        # Remove the temporary sorting field and assign default roles
        for stat in regular_speakers:
            del stat['_word_count_int']
            if 'Role' not in stat:
                stat['Role'] = 'Other'
        
        # Ensure all speaker_stats have Role assigned
        for stat in speaker_stats:
            if 'Role' not in stat:
                stat['Role'] = 'Other'
    
    speaker_view_df = pd.DataFrame(speaker_stats)
    if not speaker_view_df.empty:
        # Reorder columns
        speaker_view_df = speaker_view_df[['Role', 'Speaker', 'Speaking Duration (seconds)', 'Word Count', 'Words per Second', 'Top 10 Keywords']]
    
    # ==========================================
    # Section C: Detailed Coverage List
    # ==========================================
    print("📋 Preparing Detailed Coverage List...")
    
    # Create a copy of results for modification
    detailed_coverage_df = results_df.copy()
    
    # Enhance the Covered column with visual cues
    detailed_coverage_df['Covered'] = detailed_coverage_df['Covered'].apply(
        lambda x: f"✅ {x}" if x == 'Covered' else f"❌ {x}"
    )
    
    # Reorder columns to be more business-friendly
    business_friendly_columns = [
        'Required_Discussion_Point',
        'Covered',
        'Matched_Group',
        'Speaker',
        'Best_Matching_Variation',
        'Group_ID',
        'All_Variations_Count'
    ]
    
    # Add remaining columns (metric scores)
    remaining_columns = [col for col in detailed_coverage_df.columns if col not in business_friendly_columns]
    final_columns = business_friendly_columns + remaining_columns
    
    # Reorder the DataFrame
    detailed_coverage_df = detailed_coverage_df[final_columns]
    
    # Return the complete consolidated data for reference
    return {
        'kpis': kpis_df,
        'speaker_view': speaker_view_df,
        'detailed_coverage': detailed_coverage_df
    }

def write_executive_summary_content(ws, kpis_df, speaker_view_df, detailed_coverage_df):
    """Write Executive Summary content to a worksheet"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    import string
    
    current_row = 1
    
    # Define styles
    header_font = Font(bold=True, size=12)
    table_header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Section A: Overall Performance KPIs
    ws[f'A{current_row}'] = "A. Overall Performance KPIs"
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    # Write KPIs data
    for idx, row in kpis_df.iterrows():
        ws[f'A{current_row}'] = row['Metric']
        ws[f'B{current_row}'] = row['Value']
        ws[f'A{current_row}'].font = table_header_font
        current_row += 1
    
    current_row += 2
    
    # Section B: Key Insights - Speaker View
    ws[f'A{current_row}'] = "B. Key Insights - Speaker View"
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    if not speaker_view_df.empty:
        # Write headers
        for col_idx, col_name in enumerate(speaker_view_df.columns, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = col_name
            cell.font = table_header_font
            cell.fill = header_fill
            cell.border = border
        current_row += 1
        
        # Write data
        for idx, row in speaker_view_df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.border = border
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No speaker data available"
        current_row += 1
    
    current_row += 2
    
    # Section C: Detailed Coverage List
    ws[f'A{current_row}'] = "C. Detailed Coverage List"
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    if not detailed_coverage_df.empty:
        # Write headers
        for col_idx, col_name in enumerate(detailed_coverage_df.columns, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = col_name
            cell.font = table_header_font
            cell.fill = header_fill
            cell.border = border
        
        detail_header_row = current_row
        current_row += 1
        
        # Write data
        for idx, row in detailed_coverage_df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.border = border
            current_row += 1
        
        # Enable AutoFilter on the detailed coverage table
        detail_end_col = len(detailed_coverage_df.columns)
        detail_end_row = current_row - 1
        detail_end_col_letter = chr(ord('A') + detail_end_col - 1)
        ws.auto_filter.ref = f"A{detail_header_row}:{detail_end_col_letter}{detail_end_row}"
    else:
        ws[f'A{current_row}'] = "No coverage data available"
        current_row += 1
    
    # Adjust column widths for readability
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long content
        ws.column_dimensions[column_letter].width = adjusted_width

# ==========================================================
# Main Routine - Four-Step Early Separation Strategy
# ==========================================================
# ==========================================================
# OLD MAIN FUNCTION REMOVED
# ==========================================================
# The previous main function that used global variables (SCRIPT_SHEET_NAME, 
# CALL_TEXT_FILE_PATH, etc.) has been removed and replaced with the modular 
# run_analysis function approach for batch processing compatibility.
# ==========================================================

# ==========================================================
# ==========================================================
def run_analysis(call_file_path, script_file_path, script_sheet_name, output_folder, call_type=None, language=None):
    """
    Core analysis function implementing the four-step strategy:
    1. Early Separation: Split call_df into analysis_df and system_audio_df
    2. Independent Processing: Only process analysis_df for grouping
    3. Focused Analysis: check_coverage only analyzes human dialogue
    4. Late Merge & Reporting: Combine results for final reports
    
    Args:
        call_file_path: Path to the call text file
        script_file_path: Path to the script file
        script_sheet_name: Name of the script sheet
        output_folder: Folder to save output files
        call_type: Type of call ('SQCCB' or 'Sales Call'), if None will auto-detect
        language: Language code ('MAN', 'ENG'), if None will auto-detect
        
    Returns:
        dict: Results dictionary with status, output_file, and coverage_rate
    """
    # Start tracing memory allocations and overall timer
    tracemalloc.start()
    overall_start = print_checkpoint(1, "Program startup, loading data files")
    
    # 🔍 通话类型检测逻辑（单一事实来源）
    if call_type is None:
        # Fallback: 从文件名检测（仅在直接调用时使用）
        filename = os.path.basename(call_file_path).upper()
        if "SQCCB" in filename:
            call_type = "SQCCB"
        else:
            call_type = "Sales Call"
        print(f"⚠️  Fallback detection - 通话类型: {call_type}")
    else:
        print(f"✅ 接收到的通话类型: {call_type}")
    
    print(f"📁 文件名: {os.path.basename(call_file_path)}")
    
    # Check if files exist
    if not os.path.exists(call_file_path):
        print(f"❌ Error: Call text file '{call_file_path}' not found.")
        return {'status': 'ERROR', 'error': f'Call file not found: {call_file_path}', 'output_file': None, 'coverage_rate': 0.0}
        
    if not os.path.exists(script_file_path):
        print(f"❌ Error: Script file '{script_file_path}' not found.")
        return {'status': 'ERROR', 'error': f'Script file not found: {script_file_path}', 'output_file': None, 'coverage_rate': 0.0}
    
    try:
        step1_start = time.time()
        
        # Load call text file (CSV format)
        print(f"📁 Loading call text from: {call_file_path}")
        if call_file_path.endswith('.csv'):
            call_df = pd.read_csv(call_file_path)
            print(f"   Loading CSV file directly")
        elif call_file_path.endswith('.xlsx') or call_file_path.endswith('.xls'):
            xl_call = pd.ExcelFile(call_file_path, engine='openpyxl')
            call_df = xl_call.parse(xl_call.sheet_names[0])  # Use first sheet
            print(f"   Using sheet: {xl_call.sheet_names[0]}")
        else:
            raise ValueError(f"Unsupported file format. Please use .csv, .xlsx, or .xls files.")
        
        # Load script file
        if script_file_path.endswith('.xlsx'):
            xl_script = pd.ExcelFile(script_file_path, engine='openpyxl')
            script_df = xl_script.parse(script_sheet_name)
        else:
            script_df = pd.read_csv(script_file_path)
            
        print_checkpoint(2, f"Excel文件加载完成 ({len(call_df):,} 通话行, {len(script_df)} 脚本行)", step1_start)
    except Exception as e:
        print(f"❌ 读取Excel文件错误: {e}")
        return {'status': 'ERROR', 'error': str(e), 'output_file': None, 'coverage_rate': 0.0}

    # Initialize the checker
    step2_start = time.time()
    checker = SbertCallCoverageChecker(model_path=SBERT_MODEL_PATH)
    
    # 🔍 语言检测逻辑（单一事实来源）
    if language is None:
        # Fallback: 从文件名检测语言（仅在直接调用时使用）
        filename = os.path.basename(call_file_path).upper()
        if "_M" in filename:
            language = "MAN"
        elif "_E" in filename:
            language = "ENG"
        else:
            language = "MAN"  # 默认普通话
        print(f"⚠️  Fallback detection - 语言: {language}")
    else:
        print(f"✅ 接收到的语言: {language}")
    
    # 设置checker的语言
    checker.current_language = language
    
    # Load call-specific weights for all call types (unified processing)
    # This function is intelligent enough to handle both Sales Call and SQCCB based on script_sheet_name
    checker.load_call_specific_weights(call_file_path, script_df, script_sheet_name)
    
    # Check if the detected language is supported
    if checker.current_language not in ['MAN', 'ENG']:
        print(f"❌ Error: Language '{checker.current_language}' is not supported in Mandarin version.")
        print(f"Mandarin version only supports MAN (Mandarin) and ENG (English).")
        return {'status': 'ERROR', 'error': f'Unsupported language: {checker.current_language}', 'output_file': None, 'coverage_rate': 0.0}
    
    print(f"✅ Language validation passed: {checker.current_language}")
    print_checkpoint(3, "初始化分析器并验证语言支持", step2_start)
    
    # STEP 1 & 2: UNIFIED PROCESSING - Build grouped lines with call_type-aware system recording detection
    step3_start = time.time()
    print(f"\n📍 统一处理 - 使用 {call_type} 模式进行分组...")
    
    # Build grouped lines using the unified function that handles system recording detection internally
    grouped_lines, original_to_group, updated_call_df = checker.build_grouped_lines(call_df, call_type=call_type)
    
    # Create analysis_df and system_audio_df for later use (required by create_call_text_analysis_view_with_separation)
    system_recording_flags = checker.detect_system_audio_recordings(call_df)
    analysis_df = call_df[~system_recording_flags].copy()
    system_audio_df = call_df[system_recording_flags].copy()
    
    print(f"✅ 分组处理完成:")
    print(f"   - 创建了 {len(grouped_lines)} 个组")
    print_checkpoint(4, f"创建并保存分组通话数据 ({len(grouped_lines)} 个分组)", step3_start)
    
    # STEP 3: FOCUSED ANALYSIS - Coverage analysis on human dialogue only
    step5_start = time.time()
    print("\n📍 步骤3: 专注分析 - 对人类对话进行覆盖率分析...")
    
    # Check coverage using ONLY the human dialogue groups
    results = checker.check_coverage(updated_call_df, script_df, grouped_lines, threshold=0.4)
    
    print(f"✅ 覆盖率分析完成:")
    print(f"   - 分析了 {len(results)} 个讨论点")
    print(f"   - 使用了 {len(grouped_lines)} 个人类对话组")
    print_checkpoint(5, f"完成覆盖分析 ({len(results)} 个要点)", step5_start)
    
    # STEP 4: LATE MERGE & REPORTING - Combine results for final reports
    step6_start = time.time()
    print("\n📍 步骤4: 后期合并与报告 - 创建综合报告...")
    
    # Create call text analysis with proper merging
    call_text_analysis = checker.create_call_text_analysis_view_with_separation(
        grouped_lines, system_audio_df, script_df, threshold=0.4
    )
    
    # Create sentence level output
    sentence_level_output = checker.create_sentence_level_output(updated_call_df, original_to_group)
    print_checkpoint(6, f"生成最终分析视图", step6_start)
    
    # Generate Reports with Executive Summary
    step7_start = time.time()
    print("\n📊 生成包含Executive Summary的最终报告...")
    
    # Generate unique output filename based on input filename and save to output_folder
    input_filename = os.path.basename(call_file_path)
    if input_filename.endswith('.csv'):
        # Remove .wav if present, then replace .csv with .xlsx
        if '.wav' in input_filename:
            output_filename = input_filename.replace('.wav.csv', '.xlsx')
        else:
            output_filename = input_filename.replace('.csv', '.xlsx')
    elif input_filename.endswith('.xlsx') or input_filename.endswith('.xls'):
        # Remove .wav if present, then add _result
        if '.wav' in input_filename:
            output_filename = input_filename.replace('.wav.xlsx', '.xlsx').replace('.wav.xls', '.xlsx')
        else:
            output_filename = input_filename.replace('.xlsx', '_result.xlsx').replace('.xls', '_result.xlsx')
    else:
        # Remove .wav if present, then add _result
        if '.wav' in input_filename:
            output_filename = input_filename.replace('.wav', '') + '_result.xlsx'
        else:
            output_filename = f"{input_filename}_result.xlsx"
    
    # Create full output path
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, output_filename)
    
    print(f"📝 Output file will be: {output_file}")
    try:
        # Generate the Executive Summary data first
        summary_data = generate_executive_summary_report(
            results, updated_call_df, script_df, grouped_lines, None  # Don't save to separate file
        )
        
        # Save to the main output file with Executive Summary as first sheet
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # First, create Executive Summary sheet manually
            wb = writer.book
            ws = wb.create_sheet("Executive Summary", 0)
            write_executive_summary_content(ws, summary_data['kpis'], summary_data['speaker_view'], summary_data['detailed_coverage'])
            
            # Then write other sheets
            call_text_analysis.to_excel(writer, sheet_name='Call Text Analysis', index=False)
            sentence_level_output.to_excel(writer, sheet_name='Sentence Level Analysis', index=False)
            
            # Remove the default 'Sheet' if it exists
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # Verify file creation
        if os.path.exists(output_file):
            file_size = os.path.getsize(output_file)
            print(f"✅ Excel文件成功创建: {output_file} ({file_size:,} bytes)")
        else:
            print(f"❌ Excel文件未创建: {output_file}")
            return {'status': 'ERROR', 'error': 'File creation failed', 'output_file': None, 'coverage_rate': 0.0}
            
    except Exception as e:
        print(f"❌ 生成报告错误: {e}")
        print("尝试保存备份CSV文件...")
        backup_folder = os.path.join(output_folder, 'backup')
        os.makedirs(backup_folder, exist_ok=True)
        results.to_csv(os.path.join(backup_folder, f'coverage_analysis_backup_{os.path.splitext(output_filename)[0]}.csv'), index=False, encoding='utf-8-sig')
        call_text_analysis.to_csv(os.path.join(backup_folder, f'call_text_analysis_backup_{os.path.splitext(output_filename)[0]}.csv'), index=False, encoding='utf-8-sig')
        sentence_level_output.to_csv(os.path.join(backup_folder, f'sentence_level_analysis_backup_{os.path.splitext(output_filename)[0]}.csv'), index=False, encoding='utf-8-sig')
        return {'status': 'ERROR', 'error': str(e), 'output_file': None, 'coverage_rate': 0.0}
    
    # Create grouped call DataFrame for reference
    grouped_call_df = pd.DataFrame(grouped_lines)
    reference_file = os.path.join(output_folder, f"grouped_call_data_{os.path.splitext(output_filename)[0]}.xlsx")
    grouped_call_df.to_excel(reference_file, index=False)
    
    print_checkpoint(7, f"生成包含执行摘要的报告 ({output_file})", step7_start)
    
    # Calculate final metrics
    total_time = time.time() - overall_start
    covered_count = len(results[results['Covered'] == 'Covered'])
    total_points = len(results)
    coverage_rate = (covered_count / total_points * 100) if total_points > 0 else 0.0
    
    print(f"\n🎉 四步早期分离策略分析完成! 总时间: {total_time:.1f}s")
    print(f"   • STEP 1: Early Separation - System recording detection")
    print(f"   • STEP 2: Independent Processing - Human dialogue grouping")
    print(f"   • STEP 3: Focused Analysis - Coverage analysis on human dialogue only")
    print(f"   • STEP 4: Late Merge & Reporting - Comprehensive report generation")
    print(f"   📋 Files created:")
    print(f"   • {output_file}: Executive Summary with late merge strategy")
    print(f"   • {reference_file}: Pure human dialogue groups")
    
    print(f"\n📊 Coverage Summary:")
    print(f"   • Coverage Rate: {coverage_rate:.1f}% ({covered_count}/{total_points} points)")
    print(f"   • Total groups created: {len(grouped_lines)}")
    print(f"   • Call type: {call_type}")

    print(f"\n📊 Output Files Created:")
    print(f"  • {output_file} (3 sheets with Executive Summary):")
    print(f"    - Executive Summary: Consolidated business-friendly report with KPIs, Speaker View & Detailed Coverage")
    print(f"    - Call Text Analysis: {len(call_text_analysis)} call text segments analysis")
    print(f"    - Sentence Level Analysis: {len(sentence_level_output)} sentence-level analysis")
    print(f"  • {reference_file}: {len(grouped_call_df)} grouped call data reference")
    print(f"\n🎉 Executive Summary Features:")
    print(f"  • 📈 Overall Performance KPIs: Total duration, coverage rate, risk points")
    print(f"  • 🎤 Speaker View: Role identification (Sales/Customer/System), duration, keywords")
    print(f"  • 📋 Detailed Coverage List: Business-friendly format with ✅/❌ indicators")
    print(f"  • 💼 Professional Excel formatting: Bold headers, borders, auto-filter")
    print(f"  • 🔄 Dynamic filename: {input_filename} → {output_filename}")

    # Final checkpoint and cleanup
    overall_time = print_checkpoint("FINAL", "All processing completed", overall_start)
    
    # Print memory usage
    current, peak = tracemalloc.get_traced_memory()
    print(f"\n💾 Memory usage summary:")
    print(f"  - Current memory usage: {current / 10**6:.2f}MB")
    print(f"  - Peak memory usage: {peak / 10**6:.2f}MB")

    # Stop tracing memory allocations
    tracemalloc.stop()
    
    print(f"\n🎉 Program execution completed! Total time: {overall_time - overall_start:.2f}s")
    
    # Extract speaker word counts from summary_data
    sales_word_count = 0
    customer_word_count = 0
    
    if summary_data and 'speaker_view' in summary_data:
        speaker_view_df = summary_data['speaker_view']
        if not speaker_view_df.empty:
            for _, row in speaker_view_df.iterrows():
                if row.get('Role') == 'Sales' and pd.notna(row.get('Word Count')) and str(row.get('Word Count')).isdigit():
                    sales_word_count = int(row['Word Count'])
                elif row.get('Role') == 'Customer' and pd.notna(row.get('Word Count')) and str(row.get('Word Count')).isdigit():
                    customer_word_count = int(row['Word Count'])
    
    # Convert total_call_duration to seconds (numeric)
    total_call_duration_seconds = 0
    if summary_data and 'kpis' in summary_data:
        kpis_df = summary_data['kpis']
        duration_row = kpis_df[kpis_df['Metric'] == 'Total Call Duration']
        if not duration_row.empty:
            duration_value = duration_row.iloc[0]['Value']
            if pd.notna(duration_value) and str(duration_value) != 'N/A':
                try:
                    # Parse time string to seconds
                    duration_str = str(duration_value)
                    if ':' in duration_str:
                        parts = duration_str.split(':')
                        if len(parts) == 2:  # MM:SS
                            total_call_duration_seconds = int(parts[0]) * 60 + int(parts[1])
                        elif len(parts) == 3:  # HH:MM:SS
                            total_call_duration_seconds = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                    else:
                        # Direct numeric value
                        total_call_duration_seconds = float(duration_str)
                except (ValueError, TypeError):
                    total_call_duration_seconds = 0
    
    # Return enhanced results dictionary
    return {
        'status': 'SUCCESS',
        'output_file': output_file,
        'coverage_rate': coverage_rate,
        'total_points': total_points,
        'covered_points': covered_count,
        'total_call_duration': total_call_duration_seconds,
        'sales_word_count': sales_word_count,
        'customer_word_count': customer_word_count,
        'processing_time': total_time,
        'call_type': call_type,
        'language': checker.current_language
    }


def main():
    """
    Main function for standalone testing using default configuration.
    Note: When used standalone, you need to modify the paths below or use run_batch_analysis.py
    """
    # Default paths for standalone testing - modify these as needed
    default_call_file = "call_text_sample_M.wav.csv"
    default_script_file = "Scripts.xlsx" 
    default_script_sheet = "Script"
    default_output_folder = "."
    
    print("⚠️  Running in standalone mode with default paths.")
    print("For batch processing, use run_batch_analysis.py instead.")
    
    return run_analysis(default_call_file, default_script_file, default_script_sheet, default_output_folder)


if __name__ == "__main__":
    main()
